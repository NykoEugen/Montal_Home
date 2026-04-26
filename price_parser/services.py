import re
import json
import logging
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin, urlparse

import csv
import requests
from bs4 import BeautifulSoup
from django.db import close_old_connections
from django.db.utils import InterfaceError, OperationalError
from django.utils import timezone

from .models import (
    GoogleSheetConfig,
    PriceUpdateLog,
    FurniturePriceCellMapping,
    SupplierFeedConfig,
    SupplierFeedUpdateLog,
    SupplierWebConfig,
    SupplierWebUpdateLog,
)
from furniture.models import Furniture, FurnitureSizeVariant

logger = logging.getLogger(__name__)

DEFAULT_FEED_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": "application/xml,text/xml;q=0.9,*/*;q=0.8",
}


class GoogleSheetsPriceUpdater:
    """Service for updating furniture prices from Google Sheets."""
    
    def __init__(self, config: GoogleSheetConfig):
        self.config = config
        self.log = None
    
    def test_parse(self) -> Dict:
        """Test fetching sheet data without updating prices."""
        try:
            data = self._fetch_sheet_data()
            if not data:
                return {'success': False, 'error': 'Не вдалося отримати дані з таблиці'}
            
            # Count active cell mappings
            cell_mappings_count = FurniturePriceCellMapping.objects.filter(
                config=self.config,
                is_active=True
            ).count()
            
            return {
                'success': True,
                'data': data[:10],  # Return first 10 rows for preview
                'count': cell_mappings_count,
                'message': f'Sheet loaded successfully. Found {cell_mappings_count} active cell mappings.'
            }
        except Exception as e:
            logger.error(f"Error testing sheet access for config {self.config.name}: {str(e)}")
            return {'success': False, 'error': str(e)}
    
    def update_prices(self) -> Dict:
        """Update furniture prices from Google Sheets using direct cell mappings."""
        start_time = timezone.now()
        
        # Create log entry
        self.log = PriceUpdateLog.objects.create(
            config=self.config,
            status='success',
            started_at=start_time
        )
        
        try:
            # Fetch data from Google Sheets
            data = self._fetch_sheet_data()
            if not data:
                self._update_log_error("Не вдалося отримати дані з таблиці")
                return {'success': False, 'error': 'Не вдалося отримати дані з таблиці'}
            
            # Update prices using direct cell mappings only
            updated_count, processed_count = self._update_prices_from_cell_mappings(data)
            
            # Update log
            self.log.items_processed = processed_count
            self.log.items_updated = updated_count
            self.log.completed_at = timezone.now()
            self.log.log_details = f"Оновлено {updated_count} товарів з {processed_count} комірок"
            self.log.save()
            
            return {
                'success': True,
                'updated_count': updated_count,
                'processed_count': processed_count,
            }
            
        except Exception as e:
            error_msg = f"Помилка оновлення цін: {str(e)}"
            logger.error(error_msg)
            self._update_log_error(error_msg)
            return {'success': False, 'error': str(e)}
    
    def _fetch_sheet_data(self) -> Optional[List[List]]:
        """Fetch data from Google Sheets or XLSX file."""
        try:
            if self.config.xlsx_file:
                # Handle XLSX file
                return self._fetch_xlsx_data()
            else:
                # Handle Google Sheets
                return self._fetch_google_sheets_data()
        except Exception as e:
            logger.error(f"Error fetching data: {str(e)}")
            return None
    
    def _fetch_google_sheets_data(self) -> Optional[List[List]]:
        """Fetch data from Google Sheets using CSV export with GViz fallback."""
        last_error: Optional[Exception] = None

        try:
            return self._fetch_google_sheets_via_export()
        except Exception as exc:
            last_error = exc
            logger.warning(
                "Standard sheet export failed for config %s: %s",
                self.config.name,
                exc,
            )

        try:
            return self._fetch_google_sheets_via_xlsx_export()
        except Exception as exc:
            last_error = exc
            logger.warning(
                "XLSX export fallback failed for config %s: %s",
                self.config.name,
                exc,
            )

        try:
            data = self._fetch_google_sheets_via_gviz()
            if data:
                logger.info(
                    "GViz fallback succeeded for config %s",
                    self.config.name,
                )
            return data
        except Exception as exc:
            last_error = exc
            logger.error(
                "GViz fallback failed for config %s: %s",
                self.config.name,
                exc,
            )

        if last_error:
            raise last_error
        return None

    def _fetch_google_sheets_via_export(self) -> List[List]:
        """Fetch sheet data via the export CSV endpoint."""
        csv_url = (
            f"https://docs.google.com/spreadsheets/d/{self.config.sheet_id}/export"
            f"?format=csv&gid={self._get_sheet_gid()}"
        )
        response = requests.get(csv_url, timeout=30)
        response.raise_for_status()
        csv_data = StringIO(response.text)
        return list(csv.reader(csv_data))

    def _fetch_google_sheets_via_gviz(self) -> List[List]:
        """Fetch sheet data via the GViz endpoint (works when export is blocked)."""
        base_url = f"https://docs.google.com/spreadsheets/d/{self.config.sheet_id}/gviz/tq"
        params: Dict[str, str] = {"tqx": "out:csv"}

        if self.config.sheet_gid:
            params["gid"] = self.config.sheet_gid
        elif self.config.sheet_name:
            params["sheet"] = self.config.sheet_name
        else:
            params["gid"] = self._get_sheet_gid()

        response = requests.get(base_url, params=params, timeout=30)
        response.raise_for_status()
        csv_data = StringIO(response.text)
        return list(csv.reader(csv_data))

    def _fetch_google_sheets_via_xlsx_export(self) -> List[List]:
        """Download the Google Sheet as XLSX and return matrix data."""
        from openpyxl import load_workbook  # Imported lazily

        url = f"https://docs.google.com/spreadsheets/d/{self.config.sheet_id}/export?format=xlsx&id={self.config.sheet_id}"
        params: Dict[str, str] = {}
        if self.config.sheet_gid:
            params["gid"] = self.config.sheet_gid

        response = requests.get(url, params=params or None, timeout=60)
        response.raise_for_status()

        workbook = load_workbook(BytesIO(response.content), data_only=True)

        if self.config.sheet_name and self.config.sheet_name in workbook.sheetnames:
            worksheet = workbook[self.config.sheet_name]
        else:
            worksheet = workbook.active

        return self._worksheet_to_data(worksheet)

    def _fetch_xlsx_data(self) -> Optional[List[List]]:
        """Fetch data from XLSX file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl is not installed. Please install it with: pip install openpyxl")
            return None

        try:
            workbook = load_workbook(self.config.xlsx_file.path, data_only=True)
            if self.config.sheet_name and self.config.sheet_name in workbook.sheetnames:
                worksheet = workbook[self.config.sheet_name]
            else:
                worksheet = workbook.active
            return self._worksheet_to_data(worksheet)
        except Exception as e:
            logger.error(f"Error fetching XLSX data: {str(e)}")
            return None
    
    def _get_sheet_gid(self) -> str:
        """Get the GID (Google ID) for the specific sheet."""
        try:
            # First, try to use the GID from the database
            if self.config.sheet_gid:
                logger.info(f"Using GID {self.config.sheet_gid} from database for sheet '{self.config.sheet_name}'")
                return self.config.sheet_gid
            
            # Fallback to common GIDs for different sheet names
            common_gids = {
                'Sheet1': '0',
                'Sheet2': '1234567890',
                'Sheet3': '1234567891',
                'Прайс': '0',  # Default to first sheet
                'Ціни': '0',   # Default to first sheet
                'Обідні столи': '0',
                'Стільці': '0',
                'Меблі': '0',
                'Обідні столи гурт': '1927896544',  # The sheet you're currently viewing
                'Обідні столи роздріб': '0',
                'Стільці гурт': '0',
                'Стільці роздріб': '0',
            }
            
            # Try to get from common mappings
            if self.config.sheet_name in common_gids:
                gid = common_gids[self.config.sheet_name]
                logger.info(f"Using GID {gid} for sheet '{self.config.sheet_name}' from common mappings")
                return gid
            
            # If not found, default to first sheet (gid=0)
            logger.warning(f"Sheet name '{self.config.sheet_name}' not found in common GIDs, using default (gid=0)")
            return '0'
            
        except Exception as e:
            logger.error(f"Error getting sheet GID: {str(e)}")
            return '0'  # Default to first sheet

    def _worksheet_to_data(self, worksheet) -> List[List[str]]:
        """Convert an openpyxl worksheet to a list of lists."""
        data: List[List[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            row_data: List[str] = []
            for cell_value in row:
                if cell_value is None:
                    row_data.append("")
                else:
                    row_data.append(str(cell_value))
            data.append(row_data)
        return data

    def _parse_price(self, price_str: str) -> Optional[Decimal]:
        """Parse price string into Decimal and apply multiplier."""
        try:
            # Remove common non-numeric characters
            cleaned = re.sub(r'[^\d.,]', '', price_str)
            if ',' in cleaned and '.' in cleaned:
                # Handle European format (1.234,56)
                cleaned = cleaned.replace('.', '').replace(',', '.')
            elif ',' in cleaned:
                # Handle comma as decimal separator
                cleaned = cleaned.replace(',', '.')
            
            price = Decimal(cleaned)
            
            # Apply multiplier if it's not 1.0
            if self.config.price_multiplier != Decimal('1.0'):
                price = price * self.config.price_multiplier
                logger.info(f"Applied multiplier {self.config.price_multiplier} to price {cleaned} -> {price}")
            
            return price
        except (ValueError, InvalidOperation):
            return None
    
    def _column_to_index(self, column: str) -> int:
        """Convert Excel column letter to index (A=0, B=1, etc.)."""
        # Ensure column is a string
        if not isinstance(column, str):
            raise ValueError(f"Column must be a string, got {type(column)}: {column}")
        
        result = 0
        for char in column.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result - 1
    
    def _update_prices_from_cell_mappings(self, data: List[List]) -> Tuple[int, int]:
        """Update prices using direct cell mappings."""
        updated_count = 0
        
        # Get all active cell mappings for this config
        cell_mappings = list(
            FurniturePriceCellMapping.objects.filter(
            config=self.config,
            is_active=True
        ).select_related('furniture', 'size_variant')
        )
        processed_count = len(cell_mappings)
        
        for mapping in cell_mappings:
            try:
                # Convert column letter to index
                col_idx = self._column_to_index(mapping.sheet_column)
                row_idx = mapping.sheet_row - 1  # Convert to 0-based index
                
                # Check if row and column exist in data
                if row_idx < len(data) and col_idx < len(data[row_idx]):
                    price_str = data[row_idx][col_idx].strip()
                    if price_str:
                        price = self._parse_price(price_str)
                        if price:
                            # Update the price
                            if mapping.size_variant:
                                # Update size variant price
                                mapping.size_variant.price = price
                                mapping.size_variant.save()
                                updated_count += 1
                                logger.info(f"Updated size variant price for {mapping.furniture.name}: {price}")
                            else:
                                # Update main furniture price
                                mapping.furniture.price = price
                                mapping.furniture.save()
                                updated_count += 1
                                logger.info(f"Updated furniture price for {mapping.furniture.name}: {price}")
                        else:
                            self.log.errors.append({
                                'cell': mapping.cell_reference,
                                'furniture_name': mapping.furniture.name,
                                'error': f'Не вдалося розібрати ціну: {price_str}'
                            })
                    else:
                        self.log.errors.append({
                            'cell': mapping.cell_reference,
                            'furniture_name': mapping.furniture.name,
                            'error': 'Комірка порожня'
                        })
                else:
                    self.log.errors.append({
                        'cell': mapping.cell_reference,
                        'furniture_name': mapping.furniture.name,
                        'error': f'Комірка не існує (рядок {mapping.sheet_row}, колонка {mapping.sheet_column})'
                    })
                    
            except Exception as e:
                self.log.errors.append({
                    'cell': mapping.cell_reference,
                    'furniture_name': mapping.furniture.name,
                    'error': str(e)
                })
        
        return updated_count, processed_count
    
    def _update_log_error(self, error_msg: str):
        """Update log with error information."""
        if self.log:
            self.log.status = 'error'
            self.log.completed_at = timezone.now()
            self.log.errors.append({'error': error_msg})
            self.log.save()


@dataclass
class SupplierOffer:
    offer_id: str
    name: str
    model: Optional[str]
    price: Decimal
    old_price: Optional[Decimal]
    size_width: Optional[int] = None
    size_length: Optional[int] = None


class SupplierFeedPriceUpdater:
    """Service responsible for parsing supplier XML feeds (Matrolux etc.)."""

    def __init__(self, config: SupplierFeedConfig):
        self.config = config
        self.log: Optional[SupplierFeedUpdateLog] = None
        self._furniture_index: Optional[Dict[str, Dict]] = None
        self._variant_vendor_index: Optional[Dict[str, FurnitureSizeVariant]] = None

    def test_parse(self) -> Dict:
        """Preview first offers without applying changes."""
        try:
            offers = self._fetch_offers()
        except Exception as exc:  # pragma: no cover - network faults
            logger.error("Supplier feed test failed for %s: %s", self.config.name, exc)
            return {'success': False, 'error': str(exc)}

        preview = [
            {
                'offer_id': offer.offer_id,
                'name': offer.name,
                'model': offer.model,
                'price': str(offer.price),
                'old_price': str(offer.old_price) if offer.old_price else None,
            }
            for offer in offers[:10]
        ]
        return {
            'success': True,
            'offers_total': len(offers),
            'preview': preview,
        }

    def update_prices(self) -> Dict:
        if not self.config.is_active:
            return {'success': False, 'error': 'Конфігурація неактивна'}

        self.log = SupplierFeedUpdateLog.objects.create(
            config=self.config,
            status='success',
            started_at=timezone.now(),
        )

        try:
            offers = self._fetch_offers()
            if not offers:
                self._finalize_log(errors=[{'error': 'Фід не містить пропозицій'}])
                return {'success': False, 'error': 'Фід не містить пропозицій'}

            offers_processed = len(offers)
            items_matched = 0
            items_updated = 0
            errors: List[Dict[str, str]] = []
            # De-duplicate by (furniture.pk, size_variant.pk or None) so that
            # color-variant duplicates are skipped but different sizes are each processed.
            seen_pairs: set = set()

            for offer in offers:
                size_str = (
                    f"{offer.size_width}×{offer.size_length}"
                    if offer.size_width is not None else None
                )

                # --- Крок 1: пряме зіставлення з розмірним варіантом по vendor_code ---
                # Якщо offer.model є в індексі variant_vendor_code → оновлюємо варіант
                # напряму (не шукаємо Furniture через article_code / name).
                direct_variant: Optional[FurnitureSizeVariant] = None
                if offer.model:
                    variant_index = self._get_variant_vendor_index()
                    vkey = self._normalize_article(offer.model)
                    direct_variant = variant_index.get(vkey)

                if direct_variant is not None:
                    furniture = direct_variant.furniture
                    size_variant = direct_variant
                    items_matched += 1
                    pair = (furniture.pk, size_variant.pk)
                    if pair in seen_pairs:
                        continue
                    try:
                        changed = self._apply_offer_prices(furniture, offer, size_variant=size_variant)
                    except Exception as exc:
                        errors.append({
                            'крок': '1-variant-index',
                            'offer_id': offer.offer_id,
                            'name': offer.name,
                            'vendorCode': offer.model,
                            'розмір': size_str,
                            'меблі': furniture.name,
                            'варіант_id': size_variant.pk,
                            'error': f'Не вдалося оновити ціну варіанту: {exc}',
                        })
                        continue
                    seen_pairs.add(pair)
                    if changed:
                        items_updated += 1
                    continue

                # --- Крок 2: стандартне зіставлення через Furniture ---
                try:
                    furniture = self._match_offer_to_furniture(offer)
                except Exception as exc:  # Defensive to log unexpected parsing issues
                    errors.append({
                        'крок': '2-furniture-match',
                        'offer_id': offer.offer_id,
                        'name': offer.name,
                        'vendorCode': offer.model,
                        'розмір': size_str,
                        'error': f'Помилка підбору товару: {exc}',
                    })
                    continue

                if not furniture:
                    article_hint = (
                        f'артикул "{offer.model}" (норм: "{self._normalize_article(offer.model)}")'
                        if offer.model else 'vendorCode порожній'
                    )
                    name_variants = self._generate_name_variants(offer.name)[:4]
                    errors.append({
                        'крок': '2-furniture-not-found',
                        'offer_id': offer.offer_id,
                        'name': offer.name,
                        'vendorCode': offer.model,
                        'розмір': size_str,
                        'article_info': article_hint,
                        'name_variants': name_variants,
                        'error': (
                            'Товар не знайдено в БД. '
                            f'{article_hint}. '
                            f'Спробував назви: {name_variants[:2]}. '
                            'Перевірте article_code меблів або запустіть імпорт.'
                        ),
                    })
                    continue

                items_matched += 1

                size_variant = None
                if self.config.update_size_variants:
                    size_variant = self._match_offer_to_size_variant(furniture, offer)
                    # Warn when update_size_variants=True but variant not matched.
                    # This usually means vendor_code is not set on FurnitureSizeVariant
                    # (sofas imported before vendor_code field) or dimensions mismatch.
                    if size_variant is None and offer.size_width is not None:
                        errors.append({
                            'крок': '2-variant-not-found',
                            'offer_id': offer.offer_id,
                            'name': offer.name,
                            'vendorCode': offer.model,
                            'розмір': size_str,
                            'меблі': furniture.name,
                            'меблі_id': furniture.pk,
                            'error': (
                                f'Товар "{furniture.name}" знайдено (id={furniture.pk}), '
                                f'але варіант {size_str} см не знайдено серед FurnitureSizeVariant. '
                                'Запустіть import для оновлення варіантів.'
                            ),
                        })
                        # Fallback: update furniture-level price so data is not lost
                    elif size_variant is None and self.config.update_size_variants:
                        # No size in offer at all — note that furniture.price will be updated
                        errors.append({
                            'крок': '2-no-size-in-offer',
                            'offer_id': offer.offer_id,
                            'name': offer.name,
                            'vendorCode': offer.model,
                            'меблі': furniture.name,
                            'error': (
                                f'update_size_variants=True, але розмір не знайдено в оффері '
                                f'(немає <param> і немає WxL у назві). '
                                f'Оновлюється ціна меблів "{furniture.name}" напряму. '
                                'Для диванів з розміром у назві (1.1/1.3) — встановіть vendor_code на варіантах.'
                            ),
                        })

                pair = (furniture.pk, size_variant.pk if size_variant else None)
                if pair in seen_pairs:
                    continue

                try:
                    changed = self._apply_offer_prices(furniture, offer, size_variant=size_variant)
                except Exception as exc:
                    errors.append({
                        'крок': '2-apply-prices',
                        'offer_id': offer.offer_id,
                        'name': offer.name,
                        'vendorCode': offer.model,
                        'розмір': size_str,
                        'меблі': furniture.name,
                        'error': f'Не вдалося оновити ціну: {exc}',
                    })
                    continue

                seen_pairs.add(pair)
                if changed:
                    items_updated += 1

            self._finalize_log(
                offers_processed=offers_processed,
                items_matched=items_matched,
                items_updated=items_updated,
                errors=errors,
            )

            success = not errors or items_updated > 0
            return {
                'success': success,
                'offers_processed': offers_processed,
                'items_matched': items_matched,
                'items_updated': items_updated,
                'errors': errors,
            }

        except Exception as exc:
            logger.exception("Supplier feed update failed for %s", self.config.name)
            self._finalize_log(errors=[{'error': str(exc)}], force_status='error')
            return {'success': False, 'error': str(exc)}

    # --- Internal helpers -------------------------------------------------

    def _fetch_offers(self) -> List[SupplierOffer]:
        response = requests.get(
            self.config.feed_url,
            headers=DEFAULT_FEED_HEADERS,
            timeout=60,
        )
        response.raise_for_status()
        root = ET.fromstring(response.content)

        article_tag = (self.config.article_tag_name or 'model').strip()
        prefix_parts = self.config.article_prefix_parts or 0
        size_param = (self.config.size_param_name or '').strip()

        offers: List[SupplierOffer] = []
        for offer_el in root.findall('.//offer'):
            price = self._parse_decimal(offer_el.findtext('price'))
            if price is None:
                continue
            old_price = self._parse_decimal(offer_el.findtext('oldprice'))

            raw_article = (offer_el.findtext(article_tag) or '').strip() or None
            if raw_article and prefix_parts > 0:
                parts = raw_article.split('-')
                raw_article = '-'.join(parts[:prefix_parts])

            size_width: Optional[int] = None
            size_length: Optional[int] = None
            if size_param:
                for param_el in offer_el.findall('param'):
                    if param_el.get('name') == size_param:
                        size_width, size_length = self._parse_size(param_el.text or '')
                        break

            # Fallback: extract size from offer name when not found in params.
            # Handles beds like "Ліжко Luna/Луна, Розмір ліжка 140х200".
            # Sofas without WxL in name → None, no change.
            if size_width is None:
                offer_name_for_size = (offer_el.findtext('name') or '')
                m = re.search(r'(\d+)\s*[хxХX×]\s*(\d+)', offer_name_for_size)
                if m:
                    w, l = int(m.group(1)), int(m.group(2))
                    # mm→cm: beds written in mm (e.g. 1400х2000) → 140×200 cm
                    if w > 300:
                        w = w // 10
                    if l > 300:
                        l = l // 10
                    size_width, size_length = w, l

            offer = SupplierOffer(
                offer_id=offer_el.get('id') or raw_article or (offer_el.findtext('name') or '').strip() or 'unknown',
                name=(offer_el.findtext('name') or '').strip(),
                model=raw_article,
                price=price,
                old_price=old_price,
                size_width=size_width,
                size_length=size_length,
            )
            offers.append(offer)
        return offers

    def _parse_size(self, value: str) -> Tuple[Optional[int], Optional[int]]:
        """Parse '70x190' or '70х190' (Cyrillic х) → (70, 190) as (width, length)."""
        m = re.match(r'^(\d+)[xхХX×](\d+)$', value.strip())
        if m:
            return int(m.group(1)), int(m.group(2))
        return None, None

    def _match_offer_to_size_variant(
        self, furniture: 'Furniture', offer: SupplierOffer
    ) -> Optional['FurnitureSizeVariant']:
        """Return the FurnitureSizeVariant or BedSizeVariant matching offer's size, or None."""
        if offer.size_width is None or offer.size_length is None:
            return None
        return FurnitureSizeVariant.objects.filter(
            furniture=furniture,
            width=offer.size_width,
            length=offer.size_length,
        ).first()

    def _parse_decimal(self, value: Optional[str]) -> Optional[Decimal]:
        if not value:
            return None
        cleaned = re.sub(r'[^0-9,.-]', '', value)
        if not cleaned:
            return None
        if cleaned.count(',') == 1 and cleaned.count('.') == 0:
            cleaned = cleaned.replace(',', '.')
        elif cleaned.count(',') > 0 and cleaned.count('.') > 0:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    def _match_offer_to_furniture(self, offer: SupplierOffer) -> Optional[Furniture]:
        index = self._get_furniture_index()

        if self.config.match_by_article and offer.model:
            key = self._normalize_article(offer.model)
            furniture = index['article'].get(key)
            if furniture:
                return furniture

        if not (self.config.match_by_name and offer.name):
            return None

        offer_variants = self._generate_name_variants(offer.name)
        candidates = self._collect_name_matches(offer_variants)
        if len(candidates) == 1:
            return candidates[0]

        # Fallback: partial contains search across index
        partial_candidates: List[Furniture] = []
        name_index = index['names']
        for variant in offer_variants:
            if not variant:
                continue
            for stored_name, furnitures in name_index.items():
                if variant in stored_name or stored_name in variant:
                    partial_candidates.extend(furnitures)
        unique_partial = self._deduplicate(partial_candidates)
        if len(unique_partial) == 1:
            return unique_partial[0]
        return None

    def _collect_name_matches(self, offer_variants: List[str]) -> List[Furniture]:
        index = self._get_furniture_index()['names']
        matches: List[Furniture] = []
        for variant in offer_variants:
            if not variant:
                continue
            matches.extend(index.get(variant, []))
        return self._deduplicate(matches)

    def _deduplicate(self, items: List[Furniture]) -> List[Furniture]:
        seen: set[int] = set()
        unique: List[Furniture] = []
        for furniture in items:
            pk = furniture.pk
            if pk and pk not in seen:
                seen.add(pk)
                unique.append(furniture)
        return unique

    def _generate_name_variants(self, raw_name: str) -> List[str]:
        variants = [raw_name]
        split_parts = re.split(r'[\\/|,:;()+\-]+', raw_name)
        variants.extend(split_parts)
        normalized = [self._normalize_name(part) for part in variants if part]
        # Remove duplicates while preserving order
        seen: set[str] = set()
        result: List[str] = []
        for value in normalized:
            if value and value not in seen:
                seen.add(value)
                result.append(value)
        return result

    def _normalize_name(self, value: str) -> str:
        value = value.lower()
        value = value.replace('ё', 'е').replace('ї', 'і').replace('є', 'е').replace('ґ', 'г')
        value = re.sub(r'[^a-z0-9а-яіїєґ ]+', ' ', value)
        value = re.sub(r'\s+', ' ', value)
        return value.strip()

    def _normalize_article(self, value: str) -> str:
        """Normalize article code for comparison.

        Replaces slashes, spaces and other separators with dashes so that
        'R-63-cappuccino/ black' and 'R-63-cappuccino-black' are treated
        as the same key.
        """
        value = value.strip().lower()
        value = re.sub(r'[\s/\\|,;]+', '-', value)
        value = re.sub(r'-{2,}', '-', value)
        return value.strip('-')

    def _get_furniture_index(self) -> Dict[str, Dict]:
        if self._furniture_index is not None:
            return self._furniture_index

        article_index: Dict[str, Furniture] = {}
        name_index: Dict[str, List[Furniture]] = {}
        furnitures = Furniture.objects.all().only('id', 'name', 'article_code', 'price', 'promotional_price', 'is_promotional')
        for furniture in furnitures:
            if furniture.article_code:
                key = self._normalize_article(furniture.article_code)
                article_index[key] = furniture
            for variant in self._generate_name_variants(furniture.name):
                if not variant:
                    continue
                name_index.setdefault(variant, []).append(furniture)
        self._furniture_index = {'article': article_index, 'names': name_index}
        return self._furniture_index

    def _get_variant_vendor_index(self) -> Dict[str, FurnitureSizeVariant]:
        """Lazy-built index: vendor_code → FurnitureSizeVariant."""
        if self._variant_vendor_index is not None:
            return self._variant_vendor_index

        index: Dict[str, FurnitureSizeVariant] = {}
        for variant in FurnitureSizeVariant.objects.exclude(vendor_code='').select_related('furniture'):
            key = self._normalize_article(variant.vendor_code)
            index[key] = variant
        self._variant_vendor_index = index
        return self._variant_vendor_index

    def _apply_offer_prices(
        self,
        furniture: Furniture,
        offer: SupplierOffer,
        *,
        size_variant: Optional[FurnitureSizeVariant] = None,
    ) -> bool:
        base_price, promo_price = self._resolve_prices(offer)
        if base_price is None:
            raise ValueError('Не вдалося визначити ціну')

        if size_variant is not None:
            return self._apply_size_variant_prices(size_variant, base_price, promo_price)

        updated_fields: List[str] = []
        changed = False

        if furniture.price != base_price:
            furniture.price = base_price
            updated_fields.append('price')
            changed = True

        if promo_price is not None:
            if (furniture.promotional_price != promo_price) or (not furniture.is_promotional):
                furniture.promotional_price = promo_price
                furniture.is_promotional = True
                updated_fields.extend(['promotional_price', 'is_promotional'])
                changed = True
        else:
            if furniture.is_promotional or furniture.promotional_price is not None:
                furniture.is_promotional = False
                furniture.promotional_price = None
                updated_fields.extend(['is_promotional', 'promotional_price'])
                changed = True

        if changed:
            furniture.save(update_fields=list(dict.fromkeys(updated_fields)))

        return changed

    def _apply_size_variant_prices(
        self,
        variant: FurnitureSizeVariant,
        base_price: Decimal,
        promo_price: Optional[Decimal],
    ) -> bool:
        updated_fields: List[str] = []
        changed = False

        if variant.price != base_price:
            variant.price = base_price
            updated_fields.append('price')
            changed = True

        if promo_price is not None:
            if (variant.promotional_price != promo_price) or (not variant.is_promotional):
                variant.promotional_price = promo_price
                variant.is_promotional = True
                updated_fields.extend(['promotional_price', 'is_promotional'])
                changed = True
        else:
            if variant.is_promotional or variant.promotional_price is not None:
                variant.is_promotional = False
                variant.promotional_price = None
                updated_fields.extend(['is_promotional', 'promotional_price'])
                changed = True

        if changed:
            variant.save(update_fields=list(dict.fromkeys(updated_fields)))

        return changed

    def _resolve_prices(self, offer: SupplierOffer) -> Tuple[Optional[Decimal], Optional[Decimal]]:
        price = self._apply_multiplier(offer.price)
        old_price = self._apply_multiplier(offer.old_price) if offer.old_price is not None else None
        if price is None:
            return None, None
        if old_price is not None and old_price > 0:
            return old_price, price
        return price, None

    def _apply_multiplier(self, value: Optional[Decimal]) -> Optional[Decimal]:
        if value is None:
            return None
        multiplier = self.config.price_multiplier or Decimal('1')
        try:
            result = value * multiplier
        except (TypeError, InvalidOperation):
            result = value
        return result.quantize(Decimal('0.01'))

    def _finalize_log(
        self,
        offers_processed: int = 0,
        items_matched: int = 0,
        items_updated: int = 0,
        errors: Optional[List[Dict]] = None,
        force_status: Optional[str] = None,
    ) -> None:
        if not self.log:
            return

        errors = errors or []
        status = force_status or ('success' if not errors else ('partial' if items_updated or items_matched else 'error'))
        self.log.status = status
        self.log.offers_processed = offers_processed
        self.log.items_matched = items_matched
        self.log.items_updated = items_updated
        self.log.errors = errors
        self.log.completed_at = timezone.now()

        # Build human-readable log_details with error breakdown by step type.
        step_counts: Dict[str, int] = {}
        for e in errors:
            step = e.get('крок', 'unknown')
            step_counts[step] = step_counts.get(step, 0) + 1

        step_labels = {
            '2-furniture-not-found': 'товар не знайдено в БД',
            '2-variant-not-found': 'варіант за розміром не знайдено',
            '2-no-size-in-offer': 'розмір відсутній в оффері (потрібен vendor_code)',
            '1-variant-index': 'помилка оновлення варіанту (крок 1)',
            '2-apply-prices': 'помилка запису ціни',
            '2-furniture-match': 'помилка підбору товару',
        }
        breakdown_lines = [
            f"  • {step_labels.get(step, step)}: {cnt}"
            for step, cnt in sorted(step_counts.items())
        ]
        breakdown = ("\nРозбивка помилок:\n" + "\n".join(breakdown_lines)) if breakdown_lines else ""

        self.log.log_details = (
            f"Оброблено оферів: {offers_processed}\n"
            f"Збігів знайдено: {items_matched}\n"
            f"Цін оновлено: {items_updated}\n"
            f"Помилок: {len(errors)}"
            f"{breakdown}"
        )
        self.log.save()


class SupplierWebPriceUpdater:
    """Service for updating prices by scraping supplier web pages."""

    def __init__(self, config: SupplierWebConfig):
        self.config = config
        self.log: Optional[SupplierWebUpdateLog] = None
        self._session = requests.Session()
        self._session.headers.update(
            {
                "User-Agent": DEFAULT_FEED_HEADERS["User-Agent"],
                "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.8",
            }
        )
        self._sitemap_urls_cache: Optional[List[str]] = None
        self._page_cache: Dict[str, str] = {}
        self._domain = urlparse(self.config.base_url).netloc.lower()

    def _progress(self, message: str) -> None:
        """Emit progress to both logger and stdout for long-running manual runs."""
        prefix = f"[SupplierWeb:{self.config.name}]"
        text = f"{prefix} {message}"
        logger.info(text)
        print(text, flush=True)

    def _run_db_with_retry(self, func, label: str = "db operation"):
        """Run DB operation with one reconnect+retry on dropped connection."""
        close_old_connections()
        try:
            return func()
        except (InterfaceError, OperationalError) as exc:
            self._progress(f"{label}: DB connection issue, retrying once ({exc})")
            close_old_connections()
            return func()

    def test_parse(self) -> Dict:
        """Preview collected URLs and try parsing prices from a few pages."""
        try:
            urls = self._collect_candidate_urls()
        except Exception as exc:
            logger.error("Supplier web test failed for %s: %s", self.config.name, exc)
            return {"success": False, "error": str(exc)}

        preview: List[Dict[str, str]] = []
        for url in urls[:5]:
            try:
                html = self._fetch_page_content(url)
                base_price, promo_price = self._extract_prices(html)
            except Exception:
                base_price = None
                promo_price = None
            preview.append(
                {
                    "url": url,
                    "base_price": str(base_price) if base_price is not None else "",
                    "promo_price": str(promo_price) if promo_price is not None else "",
                }
            )

        return {
            "success": True,
            "urls_total": len(urls),
            "preview": preview,
        }

    def update_prices(self) -> Dict:
        if not self.config.is_active:
            return {"success": False, "error": "Конфігурація неактивна"}

        self._progress("Start update_prices")
        self.log = self._run_db_with_retry(
            lambda: SupplierWebUpdateLog.objects.create(
                config=self.config,
                status="success",
                started_at=timezone.now(),
            ),
            label="create web update log",
        )

        try:
            candidates = self._collect_candidate_urls()
            self._progress(f"Collected candidate URLs: {len(candidates)}")
            furnitures_qs = Furniture.objects.all()
            close_old_connections()
            selected_categories = self.config.target_categories.all()
            if selected_categories.exists():
                category_names = ", ".join(selected_categories.values_list("name", flat=True))
                self._progress(f"Filter by categories: {category_names}")
                furnitures_qs = furnitures_qs.filter(
                    sub_category__category__in=selected_categories
                )

            close_old_connections()
            furnitures = list(
                furnitures_qs.only(
                    "id",
                    "name",
                    "article_code",
                    "price",
                    "promotional_price",
                    "is_promotional",
                )
            )
            self._progress(f"Furniture items to process: {len(furnitures)}")

            items_processed = len(furnitures)
            items_matched = 0
            items_updated = 0
            errors: List[Dict[str, str]] = []

            for index, furniture in enumerate(furnitures, start=1):
                close_old_connections()
                self._progress(f"[{index}/{items_processed}] Processing: {furniture.name} ({furniture.article_code})")
                try:
                    matched_url = self._find_best_url_for_furniture(furniture, candidates)
                except Exception as exc:
                    self._progress(f"[{index}/{items_processed}] URL lookup error: {exc}")
                    errors.append(
                        {
                            "furniture_id": str(furniture.id),
                            "furniture_name": furniture.name,
                            "error": f"Помилка пошуку URL: {exc}",
                        }
                    )
                    continue

                if not matched_url:
                    self._progress(f"[{index}/{items_processed}] Not found on supplier site")
                    continue

                items_matched += 1
                self._progress(f"[{index}/{items_processed}] Matched URL: {matched_url}")
                try:
                    html = self._fetch_page_content(matched_url)
                    base_price, promo_price = self._extract_prices(html)
                    if base_price is None:
                        self._progress(f"[{index}/{items_processed}] Price not found in markup")
                        errors.append(
                            {
                                "furniture_id": str(furniture.id),
                                "furniture_name": furniture.name,
                                "url": matched_url,
                                "error": "Не вдалося знайти ціну на сторінці",
                            }
                        )
                        continue

                    self._progress(
                        f"[{index}/{items_processed}] Parsed prices -> base: {base_price}, promo: {promo_price}"
                    )
                    changed = self._apply_prices(furniture, base_price, promo_price)
                    if changed:
                        items_updated += 1
                        self._progress(f"[{index}/{items_processed}] DB updated")
                    else:
                        self._progress(f"[{index}/{items_processed}] No price changes")
                except Exception as exc:
                    self._progress(f"[{index}/{items_processed}] Parse/update error: {exc}")
                    errors.append(
                        {
                            "furniture_id": str(furniture.id),
                            "furniture_name": furniture.name,
                            "url": matched_url,
                            "error": str(exc),
                        }
                    )

            self._finalize_web_log(
                items_processed=items_processed,
                items_matched=items_matched,
                items_updated=items_updated,
                errors=errors,
            )
            self._progress(
                f"Finished update_prices: processed={items_processed}, matched={items_matched}, "
                f"updated={items_updated}, errors={len(errors)}"
            )

            success = not errors or items_updated > 0
            return {
                "success": success,
                "items_processed": items_processed,
                "items_matched": items_matched,
                "items_updated": items_updated,
                "errors": errors,
            }
        except Exception as exc:
            logger.exception("Supplier web update failed for %s", self.config.name)
            self._progress(f"Fatal error: {exc}")
            self._finalize_web_log(errors=[{"error": str(exc)}], force_status="error")
            return {"success": False, "error": str(exc)}

    def _collect_candidate_urls(self) -> List[str]:
        if self._sitemap_urls_cache is not None:
            self._progress(f"Using cached URLs: {len(self._sitemap_urls_cache)}")
            return self._sitemap_urls_cache

        urls: List[str] = []
        if self.config.crawl_from_robots:
            self._progress("Collect URLs from robots/sitemap")
            urls = self._collect_urls_from_robots_and_sitemaps()

        if not urls:
            # Fallback to base URL only (search fallback will still work).
            self._progress("No sitemap URLs found, fallback to base URL")
            urls = [self.config.base_url]

        normalized: List[str] = []
        seen: set[str] = set()
        for url in urls:
            if not self._is_same_domain(url):
                continue
            cleaned = url.split("#")[0]
            if cleaned not in seen:
                seen.add(cleaned)
                normalized.append(cleaned)
            if len(normalized) >= self.config.max_urls_to_scan:
                break

        self._progress(f"Prepared normalized candidate URLs: {len(normalized)}")
        self._sitemap_urls_cache = normalized
        return normalized

    def _collect_urls_from_robots_and_sitemaps(self) -> List[str]:
        base = self.config.base_url.rstrip("/")
        robots_url = f"{base}/robots.txt"
        self._progress(f"Fetch robots.txt: {robots_url}")
        sitemaps = self._extract_sitemaps_from_robots(robots_url)
        if not sitemaps:
            sitemaps = [f"{base}/sitemap.xml"]
            self._progress("No sitemap in robots.txt, fallback to /sitemap.xml")
        else:
            self._progress(f"Sitemaps discovered: {len(sitemaps)}")

        urls: List[str] = []
        for sitemap_url in sitemaps:
            try:
                self._progress(f"Parse sitemap: {sitemap_url}")
                urls.extend(self._parse_sitemap_recursive(sitemap_url))
            except Exception:
                logger.warning("Failed to parse sitemap %s", sitemap_url, exc_info=True)
            if len(urls) >= self.config.max_urls_to_scan:
                break
        self._progress(f"URLs collected from sitemaps: {len(urls[: self.config.max_urls_to_scan])}")
        return urls[: self.config.max_urls_to_scan]

    def _extract_sitemaps_from_robots(self, robots_url: str) -> List[str]:
        timeout = max(5, int(self.config.request_timeout))
        sitemaps: List[str] = []
        try:
            response = self._session.get(robots_url, timeout=timeout)
            if response.status_code >= 400:
                return []
            for line in response.text.splitlines():
                if line.lower().startswith("sitemap:"):
                    value = line.split(":", 1)[1].strip()
                    if value:
                        sitemaps.append(value)
        except Exception:
            return []
        return sitemaps

    def _parse_sitemap_recursive(self, sitemap_url: str) -> List[str]:
        timeout = max(5, int(self.config.request_timeout))
        response = self._session.get(sitemap_url, timeout=timeout)
        response.raise_for_status()
        root = ET.fromstring(response.content)
        ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

        if root.tag.endswith("sitemapindex"):
            urls: List[str] = []
            for sitemap_el in root.findall("sm:sitemap", ns):
                loc = sitemap_el.findtext("sm:loc", default="", namespaces=ns).strip()
                if not loc:
                    continue
                urls.extend(self._parse_sitemap_recursive(loc))
                if len(urls) >= self.config.max_urls_to_scan:
                    break
            return urls

        urls = []
        for url_el in root.findall("sm:url", ns):
            loc = url_el.findtext("sm:loc", default="", namespaces=ns).strip()
            if loc:
                urls.append(loc)
            if len(urls) >= self.config.max_urls_to_scan:
                break
        return urls

    def _find_best_url_for_furniture(self, furniture: Furniture, urls: List[str]) -> Optional[str]:
        article = self._normalize_text(furniture.article_code)
        name = self._normalize_text(furniture.name)

        article_matches: List[str] = []
        name_matches: List[str] = []
        for url in urls:
            norm_url = self._normalize_text(url)
            if self.config.match_by_article and article and article in norm_url:
                article_matches.append(url)
            if self.config.match_by_name and name and self._name_matches_in_url(name, norm_url):
                name_matches.append(url)

        for candidate in article_matches[:10]:
            if self._page_seems_related(candidate, furniture):
                return candidate

        for candidate in name_matches[:10]:
            if self._page_seems_related(candidate, furniture):
                return candidate

        query_terms = []
        if self.config.match_by_article and furniture.article_code:
            query_terms.append(furniture.article_code)
        if self.config.match_by_name and furniture.name:
            query_terms.append(furniture.name)

        for query in query_terms:
            search_url = self._build_search_url(query)
            if not search_url:
                continue
            self._progress(f"Search fallback for '{query}' -> {search_url}")
            found = self._extract_first_product_link_from_search(search_url, furniture)
            if found:
                return found

        return None

    def _build_search_url(self, query: str) -> Optional[str]:
        template = (self.config.search_path_template or "").strip()
        if not template:
            return None
        # Keep readable search text in URL template; requests/browser layer will
        # handle transport encoding of non-ASCII safely.
        search_query = self._prepare_search_query(query)
        path = template.replace("{query}", search_query)
        return urljoin(self.config.base_url, path)

    def _prepare_search_query(self, query: str) -> str:
        """Normalize search text to avoid accidental duplicated/dirty queries."""
        cleaned = re.sub(r"\s+", " ", (query or "")).strip()
        if not cleaned:
            return ""

        # Guard against accidental full duplication: "abcabc" -> "abc".
        if len(cleaned) % 2 == 0:
            half = len(cleaned) // 2
            if cleaned[:half] == cleaned[half:]:
                cleaned = cleaned[:half].strip()

        return cleaned

    def _extract_first_product_link_from_search(self, search_url: str, furniture: Furniture) -> Optional[str]:
        html = self._fetch_page_content(search_url)
        soup = BeautifulSoup(html, "html.parser")
        links = soup.select("a[href]")
        for link in links:
            href = (link.get("href") or "").strip()
            if not href:
                continue
            url = urljoin(search_url, href)
            if not self._is_same_domain(url):
                continue
            text = self._normalize_text(link.get_text(" ", strip=True))
            if furniture.article_code and self._normalize_text(furniture.article_code) in text:
                return url
            if self._name_matches_in_url(self._normalize_text(furniture.name), text):
                return url
        return None

    def _page_seems_related(self, url: str, furniture: Furniture) -> bool:
        try:
            html = self._fetch_page_content(url)
        except Exception:
            return False

        text_norm = self._normalize_text(html)
        if self.config.match_by_article and furniture.article_code:
            if self._normalize_text(furniture.article_code) in text_norm:
                return True
        if self.config.match_by_name and furniture.name:
            if self._name_matches_in_url(self._normalize_text(furniture.name), text_norm):
                return True
        return False

    def _fetch_page_content(self, url: str) -> str:
        if url in self._page_cache:
            return self._page_cache[url]

        timeout = max(5, int(self.config.request_timeout))
        if self.config.use_selenium:
            try:
                html = self._fetch_with_selenium(url)
                self._page_cache[url] = html
                return html
            except Exception:
                logger.warning("Selenium fetch failed for %s, fallback to requests", url, exc_info=True)

        response = self._session.get(url, timeout=timeout)
        response.raise_for_status()
        self._page_cache[url] = response.text
        return response.text

    def _fetch_with_selenium(self, url: str) -> str:
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.options import Options
        except ImportError as exc:
            raise RuntimeError(
                "Selenium не встановлено. Додайте пакет 'selenium' у середовище або вимкніть use_selenium."
            ) from exc

        options = Options()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1400,1200")

        driver = webdriver.Chrome(options=options)
        try:
            driver.get(url)
            wait_seconds = max(1, int(self.config.selenium_wait_seconds))
            driver.implicitly_wait(wait_seconds)
            return driver.page_source
        finally:
            driver.quit()

    def _extract_prices(self, html: str) -> Tuple[Optional[Decimal], Optional[Decimal]]:
        soup = BeautifulSoup(html, "html.parser")

        # Primary selector is configurable to make parser reusable for different suppliers.
        block_selector = (self.config.price_block_selector or "").strip()
        if not block_selector:
            block_selector = "div.price.hp_price"
        price_block = soup.select_one(block_selector)
        if price_block:
            del_node = price_block.find("del")
            ins_node = price_block.find("ins")
            del_price = self._parse_decimal_from_text(del_node.get_text(" ", strip=True)) if del_node else None
            ins_price = self._parse_decimal_from_text(ins_node.get_text(" ", strip=True)) if ins_node else None

            if del_price is not None and ins_price is not None:
                return self._apply_multiplier(del_price), self._apply_multiplier(ins_price)
            if del_price is not None and ins_price is None:
                return self._apply_multiplier(del_price), None
            if del_price is None and ins_price is not None:
                # Rule from user: if no <del>, <ins> is base price.
                return self._apply_multiplier(ins_price), None
            # Fallback inside configured block: first parseable number as base.
            plain_block_price = self._parse_decimal_from_text(price_block.get_text(" ", strip=True))
            if plain_block_price is not None:
                return self._apply_multiplier(plain_block_price), None

        # Fallback to known classes.
        base_tag = soup.select_one(".autocalc-product-price")
        promo_tag = soup.select_one(".autocalc-product-special")
        base_price = self._parse_decimal_from_text(base_tag.get_text(" ", strip=True)) if base_tag else None
        promo_price = self._parse_decimal_from_text(promo_tag.get_text(" ", strip=True)) if promo_tag else None

        if base_price is None and promo_price is not None:
            return self._apply_multiplier(promo_price), None
        if base_price is not None:
            return self._apply_multiplier(base_price), self._apply_multiplier(promo_price) if promo_price else None
        return None, None

    def _parse_decimal_from_text(self, value: str) -> Optional[Decimal]:
        if not value:
            return None
        cleaned = re.sub(r"[^0-9,.\s-]", "", value)
        cleaned = cleaned.replace("\xa0", " ").replace(" ", "")
        if not cleaned:
            return None
        if cleaned.count(",") == 1 and cleaned.count(".") == 0:
            cleaned = cleaned.replace(",", ".")
        elif cleaned.count(",") > 0 and cleaned.count(".") > 0:
            cleaned = cleaned.replace(".", "").replace(",", ".")
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    def _apply_prices(self, furniture: Furniture, base_price: Decimal, promo_price: Optional[Decimal]) -> bool:
        changed = False
        updated_fields: List[str] = []

        if furniture.price != base_price:
            furniture.price = base_price
            updated_fields.append("price")
            changed = True

        if promo_price is not None and promo_price < base_price:
            if furniture.promotional_price != promo_price:
                furniture.promotional_price = promo_price
                updated_fields.append("promotional_price")
                changed = True
            if not furniture.is_promotional:
                furniture.is_promotional = True
                updated_fields.append("is_promotional")
                changed = True
        else:
            if furniture.promotional_price is not None:
                furniture.promotional_price = None
                updated_fields.append("promotional_price")
                changed = True
            if furniture.is_promotional:
                furniture.is_promotional = False
                updated_fields.append("is_promotional")
                changed = True

        if changed:
            self._run_db_with_retry(
                lambda: furniture.save(update_fields=list(dict.fromkeys(updated_fields))),
                label=f"save furniture {furniture.id}",
            )
        return changed

    def _apply_multiplier(self, value: Optional[Decimal]) -> Optional[Decimal]:
        if value is None:
            return None
        multiplier = self.config.price_multiplier or Decimal("1")
        try:
            result = value * multiplier
        except (InvalidOperation, TypeError):
            result = value
        return result.quantize(Decimal("0.01"))

    def _normalize_text(self, value: str) -> str:
        value = (value or "").lower()
        value = value.replace("ё", "е").replace("ї", "і").replace("є", "е").replace("ґ", "г")
        value = re.sub(r"[^a-z0-9а-яіїєґ]+", " ", value)
        return re.sub(r"\s+", " ", value).strip()

    def _name_matches_in_url(self, name_norm: str, target_norm: str) -> bool:
        if not name_norm or not target_norm:
            return False
        if name_norm in target_norm:
            return True
        name_parts = [part for part in name_norm.split(" ") if len(part) >= 4]
        if not name_parts:
            return False
        matches = sum(1 for part in name_parts if part in target_norm)
        return matches >= min(2, len(name_parts))

    def _is_same_domain(self, url: str) -> bool:
        try:
            parsed = urlparse(url)
        except Exception:
            return False
        if not parsed.netloc:
            return True
        return parsed.netloc.lower().endswith(self._domain)

    def _finalize_web_log(
        self,
        items_processed: int = 0,
        items_matched: int = 0,
        items_updated: int = 0,
        errors: Optional[List[Dict]] = None,
        force_status: Optional[str] = None,
    ) -> None:
        if not self.log:
            return

        errors = errors or []
        status = force_status or ("success" if not errors else ("partial" if items_updated or items_matched else "error"))
        self.log.status = status
        self.log.items_processed = items_processed
        self.log.items_matched = items_matched
        self.log.items_updated = items_updated
        self.log.errors = errors
        self.log.completed_at = timezone.now()
        self.log.log_details = (
            f"Перевірено: {items_processed}, знайдено: {items_matched}, "
            f"оновлено: {items_updated}, помилок: {len(errors)}"
        )
        try:
            self._run_db_with_retry(self.log.save, label="finalize web update log")
        except Exception as exc:
            # Never crash request on log write failure.
            self._progress(f"WARNING: failed to save web update log: {exc}")


class MatroluxeSpecScraper:
    """Scrapes bed specifications from matroluxe.ua/ua/krovati via requests + BeautifulSoup.

    Algorithm:
    1. Paginate through catalog pages, collect all product URLs.
    2. For each product page: look for <font id="product_model"> to get article code.
    3. If article code matches a Furniture in our DB → save spec table as FurnitureParameter.
    4. Ensure every new Parameter is added to SubCategory("Ліжка").allowed_params.
    """

    CATALOG_URL = "https://matroluxe.ua/ua/krovati"
    SPEC_TABLE_SELECTOR = "div.product_tab_content.tab-specification table"
    ARTICLE_LABELS = {"артикул", "код товару", "модель", "sku", "article", "код"}
    BED_SUBCATEGORY_NAME = "Ліжка"

    _UA_TRANSLIT: Dict[str, str] = {
        "а": "a", "б": "b", "в": "v", "г": "h", "ґ": "g", "д": "d", "е": "e",
        "є": "ye", "ж": "zh", "з": "z", "и": "y", "і": "i", "ї": "yi", "й": "y",
        "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
        "с": "s", "т": "t", "у": "u", "ф": "f", "х": "kh", "ц": "ts", "ч": "ch",
        "ш": "sh", "щ": "shch", "ь": "", "ю": "yu", "я": "ya",
    }

    def __init__(self, request_timeout: int = 20) -> None:
        self.request_timeout = request_timeout
        self._session = requests.Session()
        self._session.headers.update({
            "User-Agent": DEFAULT_FEED_HEADERS["User-Agent"],
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
            "Accept-Language": "uk-UA,uk;q=0.9,en;q=0.8",
        })
        self._page_cache: Dict[str, str] = {}

    def _progress(self, msg: str) -> None:
        text = f"[MatroluxeSpec] {msg}"
        logger.info(text)
        print(text, flush=True)

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    def scrape_beds(self, dry_run: bool = False, article_code: Optional[str] = None) -> Dict:
        """Main entry point. Returns summary dict."""
        from sub_categories.models import SubCategory
        from params.models import Parameter, FurnitureParameter  # noqa: F401

        try:
            bed_subcat = SubCategory.objects.get(name=self.BED_SUBCATEGORY_NAME)
        except SubCategory.DoesNotExist:
            return {"success": False, "error": f"Підкатегорія '{self.BED_SUBCATEGORY_NAME}' не знайдена"}

        beds_qs = Furniture.objects.filter(sub_category=bed_subcat)
        if article_code:
            beds_qs = beds_qs.filter(article_code=article_code)

        article_index: Dict[str, Furniture] = {
            f.article_code.lower().strip(): f for f in beds_qs if f.article_code
        }
        if not article_index:
            return {"success": False, "error": "Ліжок не знайдено за заданими критеріями"}

        self._progress(f"Beds to match: {len(article_index)}")
        self._progress("Collecting product URLs from catalog...")
        product_urls = self._collect_product_urls()
        self._progress(f"Product URLs collected: {len(product_urls)}")

        processed = matched = updated = 0
        errors: List[Dict] = []

        for idx, url in enumerate(product_urls, 1):
            self._progress(f"[{idx}/{len(product_urls)}] {url}")
            try:
                html = self._fetch_page(url)
                specs = self._extract_specs(html)
                if not specs:
                    continue

                page_article = self._find_article_on_page(html)
                if not page_article:
                    continue

                processed += 1
                furniture = article_index.get(page_article.lower().strip())
                if not furniture:
                    self._progress(f"  article '{page_article}' not in our DB — skip")
                    continue

                matched += 1
                self._progress(f"  matched → {furniture.name}")

                if dry_run:
                    self._progress(f"  DRY RUN specs: {specs}")
                else:
                    saved = self._save_specs(furniture, specs, bed_subcat)
                    self._progress(f"  saved {saved} parameters")

                updated += 1

            except Exception as exc:
                logger.exception("MatroluxeSpec: error on %s", url)
                errors.append({"url": url, "error": str(exc)})

        return {
            "success": True,
            "product_pages_visited": len(product_urls),
            "pages_with_article": processed,
            "matched": matched,
            "updated": updated,
            "errors": errors,
        }

    # ------------------------------------------------------------------ #
    # Catalog pagination                                                   #
    # ------------------------------------------------------------------ #

    def _collect_product_urls(self) -> List[str]:
        urls: List[str] = []
        seen: set = set()
        page = 1

        while True:
            catalog_url = self.CATALOG_URL if page == 1 else f"{self.CATALOG_URL}?page={page}"
            self._progress(f"Catalog page {page}: {catalog_url}")
            try:
                html = self._fetch_page(catalog_url)
            except Exception as exc:
                self._progress(f"  failed to load catalog page {page}: {exc}")
                break

            soup = BeautifulSoup(html, "html.parser")
            found_on_page = 0
            # Collect only product links from catalog cards, not nav/footer links.
            hrefs: List[str] = []
            for a in soup.select("div.product_name a[href], .catalog__item a[href]"):
                hrefs.append(a["href"])
            # Also grab data-href on product option lists (alternative source).
            for el in soup.select("[data-href]"):
                hrefs.append(el["data-href"])
            for href in hrefs:
                href = href.strip().split("?")[0].split("#")[0]
                if not href or href in seen:
                    continue
                full = href if href.startswith("http") else f"https://matroluxe.ua{href}"
                if self._is_product_url(full):
                    seen.add(full)
                    urls.append(full)
                    found_on_page += 1

            self._progress(f"  found {found_on_page} new product URLs")

            if found_on_page == 0:
                break

            has_next = bool(soup.select_one(f"a[href*='page={page + 1}']"))
            if not has_next:
                break
            page += 1

        return urls

    def _is_product_url(self, url: str) -> bool:
        if not url.startswith("https://matroluxe.ua/ua/"):
            return False
        path = url.replace("https://matroluxe.ua/ua/", "").strip("/")
        if not path or "/" in path:
            return False
        skip_prefixes = (
            "krovati", "matras", "divan", "podushk", "topper", "futon",
            "blog", "dostavka", "kontakt", "o-kompan", "vakansi", "market",
            "aktsii", "wishlist", "compare", "dlya-dyzayn", "cart", "checkout",
            "ivano", "belaya", "vinnica", "dnepr", "zhitomir", "zaporiz",
        )
        return not any(path.startswith(p) for p in skip_prefixes)

    # ------------------------------------------------------------------ #
    # Page fetching                                                        #
    # ------------------------------------------------------------------ #

    def _fetch_page(self, url: str) -> str:
        if url in self._page_cache:
            return self._page_cache[url]
        resp = self._session.get(url, timeout=self.request_timeout)
        resp.raise_for_status()
        self._page_cache[url] = resp.text
        return resp.text

    # ------------------------------------------------------------------ #
    # Spec extraction                                                      #
    # ------------------------------------------------------------------ #

    def _extract_specs(self, html: str) -> Dict[str, str]:
        soup = BeautifulSoup(html, "html.parser")
        table = soup.select_one(self.SPEC_TABLE_SELECTOR)
        if not table:
            return {}
        specs: Dict[str, str] = {}
        for row in table.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) >= 2:
                label = cells[0].get_text(strip=True)
                value = cells[1].get_text(strip=True)
                if label and value:
                    specs[label] = value
        return specs

    def _find_article_on_page(self, html: str) -> Optional[str]:
        """Extract article code from <font id="product_model">...</font>."""
        soup = BeautifulSoup(html, "html.parser")
        el = soup.find(id="product_model")
        if el:
            value = el.get_text(strip=True)
            if value:
                return value
        for label, value in self._extract_specs(html).items():
            if label.lower().strip() in self.ARTICLE_LABELS:
                return value.strip()
        return None

    # ------------------------------------------------------------------ #
    # Persistence                                                          #
    # ------------------------------------------------------------------ #

    def _label_to_key(self, label: str) -> str:
        text = label.lower().strip()
        chars = [self._UA_TRANSLIT.get(ch, ch) for ch in text]
        key = re.sub(r"[^a-z0-9]+", "-", "".join(chars)).strip("-")
        return key[:90] or "param"

    def _save_specs(self, furniture: Furniture, specs: Dict[str, str], subcat) -> int:
        from params.models import Parameter, FurnitureParameter

        saved = 0
        for label, value in specs.items():
            if label.lower().strip() in self.ARTICLE_LABELS:
                continue
            key = self._label_to_key(label)
            if not key:
                continue
            param, _ = Parameter.objects.get_or_create(key=key, defaults={"label": label})
            if not subcat.allowed_params.filter(pk=param.pk).exists():
                subcat.allowed_params.add(param)
                self._progress(f"    → new param in allowed_params: '{param.label}'")
            FurnitureParameter.objects.update_or_create(
                furniture=furniture,
                parameter=param,
                defaults={"value": value},
            )
            saved += 1
        return saved
