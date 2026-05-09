import hashlib
import logging
import re
import time
from dataclasses import dataclass, field
from decimal import Decimal
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Tuple
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from curl_cffi import requests as cffi_requests
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils.text import slugify

logger = logging.getLogger(__name__)

BASE_URL = "https://eurosof.com.ua"
REQUEST_DELAY = 1.0
IMAGE_CACHE_DIR = "supplier_cache/eurosof"
PRICE_MULTIPLIER = Decimal("1.4")
PRICE_ADDON = Decimal("2000")
XLSX_PATH = "Прайс Eurosof Б.Церква 05.05.26р.xlsx"
MAIN_SHEET = "Прайс Б.Церква 05.05.2026р."

# Column indices (0-based after skipping first None column)
# Row: [None, name, -1КАТ, Eконом, 1КАТ, 2КАТ, 3КАТ, 4КАТ, 5КАТ, size_str]
COL_NAME = 1
COL_CAT_MINUS1 = 2
COL_ECONOM = 3
COL_CAT1 = 4
COL_CAT2 = 5
COL_CAT3 = 6
COL_CAT4 = 7
COL_CAT5 = 8
COL_SIZE = 9

CATEGORY_LABELS = ["-1 категорія", "Економ", "1 категорія", "2 категорія", "3 категорія", "4 категорія", "5 категорія"]

CYRILLIC_TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "ґ": "g", "д": "d", "е": "e",
    "є": "ie", "ж": "zh", "з": "z", "и": "y", "і": "i", "ї": "yi", "й": "y",
    "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "kh", "ц": "ts", "ч": "ch",
    "ш": "sh", "щ": "shch", "ь": "", "ю": "yu", "я": "ya",
}

# Manual slug overrides for catalog names whose transliteration doesn't match site slugs
# Key: _normalize_name(catalog_name), Value: site slug (or None to skip)
MANUAL_SLUG_MAP: Dict[str, Optional[str]] = {
    "nays": "nice",    # Найс → NICE (English loanword, no phonetic link)
}


def _translit(text: str) -> str:
    # Handle digraph дж → j before single-char mapping
    result = []
    chars = list(text.lower())
    i = 0
    while i < len(chars):
        if i + 1 < len(chars) and chars[i] == "д" and chars[i + 1] == "ж":
            result.append("j")
            i += 2
        else:
            result.append(CYRILLIC_TRANSLIT.get(chars[i], chars[i]))
            i += 1
    return "".join(result)


def _normalize_name(text: str) -> str:
    """Transliterate, lowercase, remove non-alpha."""
    return re.sub(r"[^a-z0-9]", "", _translit(text))


def _fuzzy_score(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def _extract_name_from_quotes(raw: str) -> Optional[str]:
    """Extract name inside curly/angle quotes from a catalog row name."""
    m = re.search(r'[""„««]([^""»»]+)[""»»]', raw)
    return m.group(1).strip() if m else None


def _extract_modification(raw_name: str) -> str:
    """Extract modification label from the parenthetical after the product name.

    Returns empty string if the parenthetical looks like a size spec.
    """
    # Everything after the last closing quote
    after = re.sub(r'^.*[""»»]', '', raw_name).strip()
    m = re.search(r'\(([^)]+)\)', after)
    if not m:
        return ''
    text = m.group(1).strip()
    # Reject if it looks like a size spec (contains сп.м. or digit×digit)
    if re.search(r'сп\.м|sp\.m|\d+[*хx×]\d+', text, re.I):
        return ''
    return text


def _parse_sleep_size(size_str: str) -> Optional[Tuple[int, int]]:
    """'сп.м.: 140*190  габ.розм.' or '120*200' → (width, length)"""
    if not size_str:
        return None
    m = re.search(r"сп\.м\.?\s*:?\s*(\d+)\s*[*xхХ×]\s*(\d+)", size_str)
    if m:
        return int(m.group(1)), int(m.group(2))
    # Plain NxM format used in beds sheet (e.g. "90*200")
    m = re.search(r"^(\d{2,3})\s*[*xхХ×]\s*(\d{3})$", str(size_str).strip())
    if m:
        return int(m.group(1)), int(m.group(2))
    return None


def _calc_price(catalog_price: float) -> Decimal:
    return round(Decimal(str(catalog_price)) * PRICE_MULTIPLIER + PRICE_ADDON, 0)


@dataclass
class CatalogSizeRow:
    label: str                    # e.g. "140*190"
    width: int
    length: int
    prices: List[Optional[float]] # 7 prices: -1КАТ, Економ, 1КАТ, 2КАТ, 3КАТ, 4КАТ, 5КАТ


PRODUCT_TYPE_SOFA = "sofa"       # д-н, повнорозмірний диван
PRODUCT_TYPE_MINI = "mini"       # д-н міні, Назва-міні
PRODUCT_TYPE_CORNER = "corner"   # кутовий
PRODUCT_TYPE_KANAPE = "kanape"   # к-ло (крісло розкладне)
PRODUCT_TYPE_BED = "bed"         # м'яке ліжко

BED_SHEET = "Прайс Б.Церква Ліжка"

_BED_SKIP_RE = re.compile(
    r"П double|\+матрац|Спинка-бортик|Матрац в тканин|Зашити|Обшити|"
    r"Металокаркас|Металлокаркас|Ніжка регул|Стрази Swarovski|Короб розбірний|"
    r"Дно ДСП|Підйомник ліжк|Бокові бортики на|Доповнення до",
    re.I,
)

# Subcategory config for non-sofa types (created automatically)
EXTRA_SUBCATEGORIES = {
    PRODUCT_TYPE_KANAPE: {
        "name": "Крісла розкладні",
        "slug": "krisla-rozkladni-eurosof",
    },
    PRODUCT_TYPE_MINI: {
        "name": "Міні-дивани",
        "slug": "mini-divany-eurosof",
    },
}


def _product_type(raw_name: str) -> str:
    low = raw_name.lower().strip()
    if "кутовий" in low or "кут." in low or low.startswith("кут"):
        return PRODUCT_TYPE_CORNER
    if low.startswith("к-ло") or "к-ло" in low[:6]:
        return PRODUCT_TYPE_KANAPE
    # Detect mini: "міні" in the raw name (д-н "Назва-міні", "Назва міні" etc.)
    name_in_quotes_m = re.search(r'[""„««]([^""»»]+)[""»»]', raw_name)
    check_name = name_in_quotes_m.group(1) if name_in_quotes_m else raw_name
    if "міні" in check_name.lower() or "mini" in check_name.lower():
        return PRODUCT_TYPE_MINI
    return PRODUCT_TYPE_SOFA


@dataclass
class CatalogProduct:
    catalog_name: str             # extracted from quotes
    normalized_name: str          # for matching (ignores modification)
    product_type: str             # PRODUCT_TYPE_SOFA / CORNER / KANAPE / MINI
    modification: str             # human-readable, e.g. "Мдф або м'які", or ""
    modification_slug: str        # url-safe, e.g. "mdfabomyaki", or ""
    sizes: List[CatalogSizeRow]   # all size rows for this product


@dataclass
class ScrapedProduct:
    url: str
    slug: str
    h1_name: str
    normalized_name: str
    description: str
    params: Dict[str, str]
    image_urls: List[str]


# ── Catalog Parser ────────────────────────────────────────────────────────────

class EurosofCatalogParser:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path

    def _load_rows(self) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        ws = wb[MAIN_SHEET]
        return list(ws.iter_rows(values_only=True))

    def parse(self) -> List[CatalogProduct]:
        """Straight sofas + mini + kanape (everything except corner)."""
        rows = self._load_rows()
        return self._extract_products(
            rows,
            type_filter=[PRODUCT_TYPE_SOFA, PRODUCT_TYPE_MINI, PRODUCT_TYPE_KANAPE],
        )

    def parse_corner(self) -> List[CatalogProduct]:
        rows = self._load_rows()
        return self._extract_products(rows, type_filter=[PRODUCT_TYPE_CORNER])

    def _is_header_row(self, row) -> bool:
        name = row[COL_NAME]
        if not isinstance(name, str):
            return False
        stripped = name.strip()
        return (
            re.match(r"^\d+\.ВИРІБ", stripped)
            or re.match(r"^Серія\s", stripped)
            or stripped.startswith("www.")
            or re.match(r"^0[0-9]\.", stripped)
            or stripped in {"1.ВИРІБ", "2.ВИРІБ", "3.ВИРІБ", "4.ВИРІБ", "5.ВИРІБ"}
        )

    def _is_product_row(self, row) -> bool:
        """Row is a product row if it has a name AND at least one numeric price."""
        name = row[COL_NAME]
        if not isinstance(name, str) or not name.strip():
            return False
        # At least one of the price columns must be numeric
        for col in (COL_CAT_MINUS1, COL_ECONOM, COL_CAT1):
            if isinstance(row[col], (int, float)):
                return True
        return False

    def _extract_products(self, rows: list, type_filter=None) -> List[CatalogProduct]:
        """type_filter: string or list of strings; None = all types."""
        allowed = set(type_filter) if isinstance(type_filter, list) else ({type_filter} if type_filter else None)
        products: Dict[tuple, CatalogProduct] = {}
        order: List[tuple] = []

        for row in rows:
            if not row or len(row) <= COL_SIZE:
                continue
            if self._is_header_row(row):
                continue
            if not self._is_product_row(row):
                continue

            name_raw = str(row[COL_NAME]).strip()
            catalog_name = _extract_name_from_quotes(name_raw)
            if not catalog_name:
                continue

            ptype = _product_type(name_raw)
            if allowed and ptype not in allowed:
                continue

            prices = [
                row[COL_CAT_MINUS1] if isinstance(row[COL_CAT_MINUS1], (int, float)) else None,
                row[COL_ECONOM]     if isinstance(row[COL_ECONOM], (int, float)) else None,
                row[COL_CAT1]       if isinstance(row[COL_CAT1], (int, float)) else None,
                row[COL_CAT2]       if isinstance(row[COL_CAT2], (int, float)) else None,
                row[COL_CAT3]       if isinstance(row[COL_CAT3], (int, float)) else None,
                row[COL_CAT4]       if isinstance(row[COL_CAT4], (int, float)) else None,
                row[COL_CAT5]       if isinstance(row[COL_CAT5], (int, float)) else None,
            ]
            if not any(p for p in prices):
                continue

            size_str = str(row[COL_SIZE]) if row[COL_SIZE] else ""
            parsed_size = _parse_sleep_size(size_str)
            if parsed_size:
                width, length = parsed_size
                size_label = f"{width}×{length}"
            else:
                m = re.search(r"(\d{2,3})\s*(?:сп\.?м\.?|\/)", name_raw)
                if m:
                    width = int(m.group(1))
                    length = 190  # default
                else:
                    width, length = 0, 0
                size_label = f"{width}×{length}" if width else name_raw[:30]

            size_row = CatalogSizeRow(
                label=size_label,
                width=width,
                length=length,
                prices=prices,
            )

            norm = _normalize_name(catalog_name)
            modification = _extract_modification(name_raw)
            mod_slug = _normalize_name(modification)  # transliterate + strip
            key = (norm, ptype, mod_slug)
            if key not in products:
                products[key] = CatalogProduct(
                    catalog_name=catalog_name,
                    normalized_name=norm,
                    product_type=ptype,
                    modification=modification,
                    modification_slug=mod_slug,
                    sizes=[],
                )
                order.append(key)
            products[key].sizes.append(size_row)

        return [products[k] for k in order]


class EurosofBedCatalogParser(EurosofCatalogParser):
    """Parses the beds sheet; forces product_type=BED and skips add-on rows."""

    def _load_rows(self) -> list:
        import openpyxl
        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True)
        ws = wb[BED_SHEET]
        return list(ws.iter_rows(values_only=True))

    def _is_product_row(self, row) -> bool:
        name = row[COL_NAME]
        if not isinstance(name, str) or not name.strip():
            return False
        if _BED_SKIP_RE.search(name):
            return False
        return super()._is_product_row(row)

    def parse(self) -> List[CatalogProduct]:
        rows = self._load_rows()
        products = self._extract_products(rows, type_filter=None)
        for cp in products:
            cp.product_type = PRODUCT_TYPE_BED
        return products


# ── Web Scraper ───────────────────────────────────────────────────────────────

class EurosofWebScraper:
    def __init__(self):
        self.session = cffi_requests.Session(impersonate="chrome124")
        self.session.headers.update({
            "Accept-Language": "uk-UA,uk;q=0.9,en-US;q=0.8",
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
        })
        self._progress_callback = None

    def set_progress_callback(self, cb):
        self._progress_callback = cb

    def _log(self, msg: str) -> None:
        logger.info(msg)
        if self._progress_callback:
            self._progress_callback(msg)

    def _get(self, url: str) -> Optional[BeautifulSoup]:
        try:
            resp = self.session.get(url, timeout=25)
            resp.raise_for_status()
            return BeautifulSoup(resp.text, "html.parser")
        except Exception as exc:
            logger.warning("GET failed %s: %s", url, exc)
            return None

    def collect_product_urls(self, catalog_url: str) -> List[Dict]:
        """Returns list of {slug, url} dicts from a catalog page."""
        soup = self._get(catalog_url)
        if not soup:
            return []
        slugs = set()
        results = []
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if href.startswith("/products/"):
                slug = href[len("/products/"):]
                if slug and slug not in slugs:
                    slugs.add(slug)
                    results.append({"slug": slug, "url": urljoin(BASE_URL, href)})
        return results

    def scrape_product(self, url: str, slug: str) -> Optional[ScrapedProduct]:
        soup = self._get(url)
        if not soup:
            return None

        h1 = soup.find("h1")
        h1_name = h1.get_text(strip=True) if h1 else slug.upper()

        # Description
        desc_el = soup.select_one(".block__description .block__description--style")
        description = desc_el.get_text(separator="\n", strip=True) if desc_el else ""

        # Characteristics
        params: Dict[str, str] = {}
        for item in soup.select("ul.features li.features__item"):
            name_el = item.select_one(".features__name")
            wrap = item.select_one(".features__wrap")
            if not name_el or not wrap:
                continue
            divs = wrap.find_all("div", recursive=False)
            val_el = divs[-1] if divs else None
            if val_el:
                key = name_el.get_text(strip=True).rstrip(":").strip()
                val = val_el.get_text(strip=True).rstrip(";").strip()
                if key and val:
                    params[key] = val

        # Images — use resized URLs directly (originals return 404)
        image_urls: List[str] = []
        seen: set = set()
        for img in soup.find_all("img", src=True):
            src = img["src"]
            if "/files/resized/products/" not in src:
                continue
            full_url = urljoin(BASE_URL, src)
            if full_url not in seen:
                seen.add(full_url)
                image_urls.append(full_url)

        return ScrapedProduct(
            url=url,
            slug=slug,
            h1_name=h1_name,
            normalized_name=_normalize_name(h1_name),
            description=description,
            params=params,
            image_urls=image_urls,
        )

    def scrape_all(self, catalog_urls: List[str]) -> List[ScrapedProduct]:
        all_products: List[ScrapedProduct] = []
        seen_slugs: set = set()
        for cat_url in catalog_urls:
            self._log(f"Збираємо посилання з {cat_url}")
            for entry in self.collect_product_urls(cat_url):
                if entry["slug"] in seen_slugs:
                    continue
                seen_slugs.add(entry["slug"])
                time.sleep(REQUEST_DELAY)
                product = self.scrape_product(entry["url"], entry["slug"])
                if product:
                    all_products.append(product)
                    self._log(f"  ✓ {product.h1_name} ({entry['slug']})")
                else:
                    self._log(f"  ✗ Не вдалося спарсити {entry['url']}")
        return all_products


# ── Matcher ───────────────────────────────────────────────────────────────────

def match_catalog_to_scraped(
    catalog: List[CatalogProduct],
    scraped: List[ScrapedProduct],
    threshold: float = 0.65,
) -> List[Tuple[CatalogProduct, Optional[ScrapedProduct]]]:
    """Match each catalog product to the best-scoring scraped product."""
    slug_index: Dict[str, ScrapedProduct] = {sp.slug: sp for sp in scraped}
    results = []
    for cp in catalog:
        # 1. Check manual override first
        override_slug = MANUAL_SLUG_MAP.get(cp.normalized_name)
        if override_slug is not None:
            match = slug_index.get(override_slug)
            if match:
                results.append((cp, match))
                logger.debug("MANUAL MATCH: %s → %s", cp.catalog_name, match.h1_name)
                continue

        # 2. Fuzzy match against all scraped products
        best: Optional[ScrapedProduct] = None
        best_score = 0.0
        for sp in scraped:
            score = _fuzzy_score(cp.normalized_name, sp.normalized_name)
            if score > best_score:
                best_score = score
                best = sp
        if best_score >= threshold:
            results.append((cp, best))
            logger.debug("MATCH %.2f: %s → %s", best_score, cp.catalog_name, best.h1_name)
        else:
            results.append((cp, None))
            logger.warning("UNMATCHED (%.2f): %s", best_score, cp.catalog_name)
    return results


# ── Importer ──────────────────────────────────────────────────────────────────

def _image_cache_path(url: str) -> str:
    parsed_path = url.split("eurosof.com.ua")[-1]
    digest = hashlib.sha1(url.encode()).hexdigest()
    ext_m = re.search(r"\.(jpg|jpeg|png|webp)(\?|$)", url, re.I)
    ext = f".{ext_m.group(1).lower()}" if ext_m else ".jpg"
    return f"{IMAGE_CACHE_DIR}/{digest}{ext}"


def _download_image(session, url: str) -> Optional[str]:
    cache_path = _image_cache_path(url)
    if default_storage.exists(cache_path):
        return cache_path
    try:
        resp = session.get(url, timeout=30)
        resp.raise_for_status()
    except Exception as exc:
        logger.warning("Image download failed %s: %s", url, exc)
        return None
    content = ContentFile(resp.content)
    try:
        from PIL import Image as PilImage
        content.seek(0)
        with PilImage.open(content) as img:
            img.verify()
        content.seek(0)
    except Exception:
        return None
    default_storage.save(cache_path, content)
    return cache_path


def _ensure_subcategory(subcategory_name: str, subcategory_slug: str, category_name: str):
    from categories.models import Category
    from sub_categories.models import SubCategory

    category = Category.objects.filter(name=category_name).first()
    if not category:
        raise RuntimeError(f"Категорія '{category_name}' не знайдена в БД.")
    sub, _ = SubCategory.objects.get_or_create(
        slug=subcategory_slug,
        defaults={"name": subcategory_name, "category": category},
    )
    return sub


def _ensure_fabric_brand():
    from fabric_category.models import FabricBrand, FabricCategory

    brand, _ = FabricBrand.objects.get_or_create(name="Eurosof")
    # Create categories with step_index prices (0, 1, 2, 3, 4, 5, 6)
    categories = []
    for i, label in enumerate(CATEGORY_LABELS[1:], start=1):  # skip -1КАТ (it's the base)
        cat, _ = FabricCategory.objects.get_or_create(
            brand=brand,
            name=label,
            defaults={"price": Decimal(str(i))},
        )
        categories.append(cat)
    return brand, categories


def _generate_slug(name: str) -> str:
    from furniture.models import Furniture

    base = slugify(_translit(name)) or "eurosof"
    slug = base
    suffix = 1
    while Furniture.objects.filter(slug=slug).exists():
        suffix += 1
        slug = f"{base}-{suffix}"
    return slug


def _reference_step(catalog_product: CatalogProduct) -> Decimal:
    """Compute per-step price extra (in UAH after ×1.4) from reference size, rounded up."""
    import math
    rows = [r for r in catalog_product.sizes if r.prices[0] and r.prices[1]]
    if not rows:
        return Decimal("0")
    rows_sorted = sorted(rows, key=lambda r: r.width)
    ref = rows_sorted[len(rows_sorted) // 2]
    step = Decimal(str(ref.prices[1])) - Decimal(str(ref.prices[0]))
    return Decimal(str(math.ceil(float(step * PRICE_MULTIPLIER))))


class EurosofImporter:
    def __init__(self, xlsx_path: str):
        self.xlsx_path = xlsx_path
        self.session = cffi_requests.Session(impersonate="chrome124")
        self.session.headers.update({
            "Accept-Language": "uk-UA,uk;q=0.9",
            "Accept": "text/html,*/*;q=0.8",
        })
        self._progress_callback = None

    def set_progress_callback(self, cb):
        self._progress_callback = cb

    def _log(self, msg: str) -> None:
        logger.info(msg)
        if self._progress_callback:
            self._progress_callback(msg)

    def run(
        self,
        catalog_urls: List[str],
        subcategory_name: str,
        subcategory_slug: str,
        category_name: str,
        dry_run: bool = False,
        update_prices: bool = False,
        corner: bool = False,
        bed: bool = False,
    ) -> Dict:
        from furniture.models import Furniture, FurnitureImage, FurnitureSizeVariant
        from params.models import FurnitureParameter, Parameter

        # Step 1: parse catalog
        self._log("Читаємо Excel-каталог...")
        if bed:
            parser = EurosofBedCatalogParser(self.xlsx_path)
            catalog_products = parser.parse()
        elif corner:
            parser = EurosofCatalogParser(self.xlsx_path)
            catalog_products = parser.parse_corner()
        else:
            parser = EurosofCatalogParser(self.xlsx_path)
            catalog_products = parser.parse()
        self._log(f"  Знайдено {len(catalog_products)} продуктів у каталозі")

        # Step 2: scrape website
        self._log("Скрепимо сайт eurosof.com.ua...")
        web_scraper = EurosofWebScraper()
        web_scraper.set_progress_callback(self._progress_callback)
        scraped_products = web_scraper.scrape_all(catalog_urls)
        self._log(f"  Знайдено {len(scraped_products)} продуктів на сайті")

        # Step 3: match
        self._log("Матчимо каталог із сайтом...")
        matches = match_catalog_to_scraped(catalog_products, scraped_products)

        # Step 4: ensure subcategories + fabric brand
        sub_map: Dict[str, object] = {}
        if not dry_run:
            if bed:
                sub_map[PRODUCT_TYPE_BED] = _ensure_subcategory(subcategory_name, subcategory_slug, category_name)
            elif corner:
                sub_map[PRODUCT_TYPE_SOFA] = _ensure_subcategory(subcategory_name, subcategory_slug, category_name)
                sub_map[PRODUCT_TYPE_CORNER] = sub_map[PRODUCT_TYPE_SOFA]
            else:
                sub_map[PRODUCT_TYPE_SOFA] = _ensure_subcategory(subcategory_name, subcategory_slug, category_name)
                for ptype, cfg in EXTRA_SUBCATEGORIES.items():
                    sub_map[ptype] = _ensure_subcategory(cfg["name"], cfg["slug"], category_name)
            fabric_brand, fabric_cats = _ensure_fabric_brand()

        stats = {"created": 0, "updated": 0, "skipped": 0, "unmatched": 0, "errors": []}

        for cp, sp in matches:
            if sp is None:
                self._log(f"⚠ UNMATCHED: {cp.catalog_name}")
                stats["unmatched"] += 1
                continue

            self._log(f"[{cp.catalog_name}] ({cp.product_type}) → [{sp.h1_name}]")

            if dry_run:
                for sr in cp.sizes:
                    minus1 = sr.prices[0]
                    price = _calc_price(minus1) if minus1 else "N/A"
                    self._log(f"  [DRY-RUN] {cp.product_type} розмір {sr.label}: base {price} грн")
                stats["created"] += 1
                continue

            # article_code: base + type suffix (non-sofa) + modification slug
            type_suffix = "" if cp.product_type == PRODUCT_TYPE_SOFA else f"-{cp.product_type}"
            mod_suffix = f"-{cp.modification_slug[:20]}" if cp.modification_slug else ""
            article_code = f"eurosof-{sp.slug}{type_suffix}{mod_suffix}"
            target_sub = sub_map.get(cp.product_type, sub_map.get(PRODUCT_TYPE_SOFA))

            existing = Furniture.objects.filter(article_code=article_code).first()

            if existing and update_prices:
                self._update_prices(existing, cp)
                stats["updated"] += 1
                continue

            if existing:
                self._log(f"  Пропускаємо (вже є): {existing.name}")
                stats["skipped"] += 1
                continue

            # Base price from first size's -1КАТ
            first_minus1 = next((r.prices[0] for r in cp.sizes if r.prices[0]), None)
            if not first_minus1:
                self._log(f"  ✗ Немає ціни для {cp.catalog_name}")
                stats["skipped"] += 1
                continue

            try:
                self._create_product(
                    cp, sp, target_sub, fabric_brand,
                    article_code, first_minus1, Parameter, FurnitureParameter,
                )
                self._log(f"  ✓ Створено: ({article_code})")
                stats["created"] += 1
            except Exception as exc:
                self._log(f"  ✗ Помилка {cp.catalog_name}: {exc}")
                stats["errors"].append(str(exc))
                stats["skipped"] += 1

        return stats

    def _create_size_variants(self, furniture, cp: CatalogProduct) -> None:
        from furniture.models import FurnitureSizeVariant

        for sr in cp.sizes:
            minus1_price = sr.prices[0]
            if not minus1_price:
                continue
            price = _calc_price(minus1_price)
            FurnitureSizeVariant.objects.update_or_create(
                furniture=furniture,
                width=sr.width,
                length=sr.length,
                height=0,
                defaults={"price": price},
            )

    def _update_prices(self, furniture, cp: CatalogProduct) -> None:
        from furniture.models import FurnitureSizeVariant

        fabric_value = _reference_step(cp)
        furniture.fabric_value = fabric_value
        first_minus1 = next((r.prices[0] for r in cp.sizes if r.prices[0]), None)
        if first_minus1:
            furniture.price = _calc_price(first_minus1)
        furniture.save(update_fields=["price", "fabric_value"])

        for sr in cp.sizes:
            minus1 = sr.prices[0]
            if not minus1:
                continue
            FurnitureSizeVariant.objects.filter(
                furniture=furniture, width=sr.width, length=sr.length
            ).update(price=_calc_price(minus1))

        self._log(f"  ↻ Оновлено ціни: {furniture.name}")

    def _save_images(self, furniture, image_urls: List[str]) -> None:
        from furniture.models import FurnitureImage

        first_path = None
        for position, url in enumerate(image_urls):
            path = _download_image(self.session, url)
            if path:
                FurnitureImage.objects.get_or_create(
                    furniture=furniture,
                    image=path,
                    defaults={"position": position},
                )
                if first_path is None:
                    first_path = path

        # Set main image (used on catalog listing pages)
        if first_path and not furniture.image:
            furniture.image = first_path
            furniture.save(update_fields=["image"])

    def _create_product(
        self,
        cp: CatalogProduct,
        sp: ScrapedProduct,
        sub_category,
        fabric_brand,
        article_code: str,
        first_minus1: float,
        Parameter,
        FurnitureParameter,
    ) -> None:
        from django.db import transaction
        from furniture.models import Furniture

        SKIP_PARAMS = {"доступні розміри"}

        base_price = _calc_price(first_minus1)
        fabric_value = _reference_step(cp)
        mod_label = f" ({cp.modification})" if cp.modification else ""
        display_name = f"{cp.catalog_name}{mod_label}"
        # Slug derived from article_code (guaranteed unique), drop "eurosof-" prefix
        slug = _generate_slug(article_code.replace("eurosof-", "", 1))

        with transaction.atomic():
            furniture = Furniture.objects.create(
                name=display_name,
                slug=slug,
                article_code=article_code,
                description=sp.description,
                price=base_price,
                fabric_value=fabric_value,
                selected_fabric_brand=fabric_brand,
                sub_category=sub_category,
            )

            for label, val in sp.params.items():
                if label.lower() in SKIP_PARAMS:
                    continue
                key = slugify(_translit(label)) or slugify(label)
                param, _ = Parameter.objects.get_or_create(
                    key=key, defaults={"label": label}
                )
                FurnitureParameter.objects.update_or_create(
                    furniture=furniture,
                    parameter=param,
                    defaults={"value": val[:200]},
                )

            self._create_size_variants(furniture, cp)
            self._save_images(furniture, sp.image_urls)
